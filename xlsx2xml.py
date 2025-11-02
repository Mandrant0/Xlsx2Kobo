#!/usr/bin/env python3
# Ce script convertit un fichier Excel (XLSX) structuré selon des règles précises (voir README) en un ou plusieurs fichiers XML, prêts à être soumis à KoboToolbox ou ODK.
# Version modifiée : aucune balise XML ne contiendra la valeur "None" (balises vides si cellule vide).

import sys
import uuid
from lxml import etree as ET
import os, errno
import shutil
from openpyxl import load_workbook

def _has_group(book):
    return True if len(book.sheetnames) > 1 else False

def _get_col_index(sheet_index, headers, colname):
    return headers[sheet_index].index(colname)

def _gen_headers(book):
    headers = {}
    for i, sheetname in enumerate(book.sheetnames):
        sheet = book[sheetname]
        data = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers[i] = data
    return headers

def _gen_group_indices(book, sheet_index):
    colnames = _gen_headers(book)[sheet_index]
    g_parent_index_col_index = colnames.index("_parent_index")
    group_indices = {}
    sheetname = book.sheetnames[sheet_index]
    group_sheet = book[sheetname]
    parent_col = [row[g_parent_index_col_index].value for row in group_sheet.iter_rows(min_row=2)]
    for group_row, parent_row in enumerate(parent_col, start=2):
        if parent_row in group_indices:
            group_indices[parent_row].append(group_row)
        else:
            if str(parent_row) == '_parent_index':
                continue
            group_indices[parent_row] = [group_row]
    return group_indices

def _gen_group_index_list(book):
    group_index_list = []
    for s_i, sheetname in enumerate(book.sheetnames):
        if s_i == 0 or sheetname == 'IDSheet':
            continue
        groups_indices = {}
        groups_indices[sheetname] = _gen_group_indices(book, s_i)
        group_index_list.append(groups_indices)
    return group_index_list

def _parse_multi_select_data(multi_selects, header, text):
    if text == 1 or text == '1':
        name, value = header.split('/')
        if name not in multi_selects:
            multi_selects[name] = []
        multi_selects[name].append(value)

def _gen_multi_selects(parent_node, multi_selects):
    for key, value in multi_selects.items():
        colname_el = ET.SubElement(parent_node, key)
        colname_el.text = " ".join(str(v) for v in value)

def _parse_group_data(groups, header, text):
    name, value = header.split('::')
    if name not in groups:
        groups[name] = {}
    groups[name][value] = text

def _gen_groups(root, groups):
    for key, value in groups.items():
        colname_el = ET.SubElement(root, key)
        for k, v in value.items():
            el = ET.SubElement(colname_el, k)
            el.text = "" if v is None else str(v)

def _gen_xml_elements0(book, headers, row):
    data_sheet0 = book[book.sheetnames[0]]
    ID_sheet = book['IDSheet']
    KPI_ID = ID_sheet.cell(row=1, column=2).value
    KC_ID = ID_sheet.cell(row=2, column=2).value
    version_col_index = _get_col_index(0, headers, '__version__')
    version = data_sheet0.cell(row=2, column=version_col_index+1).value

    NSMAP = {"jr": 'http://openrosa.org/javarosa', "orx": 'http://openrosa.org/xforms'}
    # Utilise un nom de balise racine XML valide
    root = ET.Element("data", nsmap=NSMAP)
    root.set('id', KPI_ID)
    root.set("version", str(version) if version is not None else "")

    fhub_el = ET.SubElement(root, "formhub")
    kc_uuid_el = ET.SubElement(fhub_el, "uuid")
    kc_uuid_el.text = str(KC_ID) if KC_ID is not None else ""

    multi_selects = {}
    groups = {}
    for i, colname in enumerate(headers[0][:version_col_index]):
        text0 = data_sheet0.cell(row=row, column=i+1).value
        if "/" in str(colname):
            _parse_multi_select_data(multi_selects, colname, text0)
        elif "::" in str(colname):
            _parse_group_data(groups, colname, text0)
        else:
            colname_el = ET.SubElement(root, str(colname))
            colname_el.text = "" if text0 is None else str(text0)

    _gen_multi_selects(root, multi_selects)
    _gen_groups(root, groups)

    elems = {row: {}}
    elems[row]['root'] = root
    return elems

def _gen_group_detail(book, row, headers, data_sheet0, root):
    group_index_list = _gen_group_index_list(book)
    _index_col_index = _get_col_index(0, headers, '_index')

    for sheet_i, g_indices in enumerate(group_index_list):
        sheet_idx = sheet_i + 1
        group_sheet = book[book.sheetnames[sheet_idx]]
        group_sheetname = group_sheet.title
        _parent_index = data_sheet0.cell(row=row, column=_index_col_index+1).value
        _index1_col_index = _get_col_index(sheet_idx, headers, '_index')

        for key, group_indices in g_indices.items():
            if _parent_index in group_indices.keys():
                for group_row in group_indices[_parent_index]:
                    group_sheetname_el = ET.SubElement(root, group_sheetname)
                    multi_selects = {}
                    for group_col in range(0, _index1_col_index):
                        header = group_sheet.cell(row=1, column=group_col+1).value
                        text = group_sheet.cell(row=group_row, column=group_col+1).value
                        if "/" in str(header):
                            _parse_multi_select_data(multi_selects, header, text)
                        else:
                            column_el = ET.SubElement(group_sheetname_el, str(header))
                            column_el.text = "" if text is None else str(text)
                    _gen_multi_selects(group_sheetname_el, multi_selects)

def gen_xml(path):
    book = load_workbook(path, data_only=True)
    headers = _gen_headers(book)
    print(headers)

    data_sheet0 = book[book.sheetnames[0]]
    for row in range(2, data_sheet0.max_row + 1):
        elems = _gen_xml_elements0(book, headers, row)
        root = elems[row]['root']

        if _has_group(book):
            _gen_group_detail(book, row, headers, data_sheet0, root)

        version_el = ET.SubElement(root, "__version__")
        version_col_index = _get_col_index(0, headers, '__version__')
        version_value = data_sheet0.cell(row=row, column=version_col_index+1).value
        version_el.text = "" if version_value is None else str(version_value)

        meta_el = ET.SubElement(root, "meta")
        instance_ID_el = ET.SubElement(meta_el, "instanceID")
        _uuid_col_index = _get_col_index(0, headers, '_uuid')
        iID = data_sheet0.cell(row=row, column=_uuid_col_index+1).value
        iID = iID if iID and len(str(iID)) > 0 else str(uuid.uuid4())
        output_iID = "uuid:" + str(iID)
        instance_ID_el.text = output_iID

        tree = ET.ElementTree(root)
        temp_path = "tempfiles/"

        try:
            os.makedirs(temp_path)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

        output_fn = temp_path + str(iID) + '.xml'
        tree.write(output_fn, pretty_print=True, xml_declaration=True, encoding="utf-8")
        shutil.make_archive('xmlfiles', 'zip', 'tempfiles')

if __name__ == "__main__":
    INPUT_EXCEL_FILE = sys.argv[1]
    shutil.rmtree("tempfiles/", ignore_errors=True)
    gen_xml(INPUT_EXCEL_FILE)